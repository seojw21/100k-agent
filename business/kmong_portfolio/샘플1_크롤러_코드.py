"""
포트폴리오 데모 크롤러 — 공개 크롤링 연습 사이트(books.toscrape.com)에서
상품 데이터를 수집해 정리된 엑셀로 저장. 크몽 포트폴리오 스크린샷용.
'수집 → 검증 → 깔끔한 엑셀 정리' 흐름을 그대로 보여준다.
"""
import time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = "https://books.toscrape.com/catalogue/page-{}.html"
RATING = {"One": 1, "Two": 2, "Three": 3, "Four": 4, "Five": 5}

rows = []
for page in range(1, 4):  # 3페이지 = 60개 상품 데모
    r = requests.get(BASE.format(page), timeout=10)
    soup = BeautifulSoup(r.text, "html.parser")
    for art in soup.select("article.product_pod"):
        title = art.h3.a["title"]
        price = art.select_one(".price_color").text.replace("Â", "").strip()
        rating = RATING.get(art.select_one(".star-rating")["class"][1], 0)
        stock = art.select_one(".availability").text.strip()
        rows.append([title, price, rating, stock])
    time.sleep(0.4)  # 서버 예의: 요청 간격

# --- 검증: 빠진 값 체크 ---
missing = sum(1 for row in rows if not all(str(c).strip() for c in row))
print(f"수집 완료: {len(rows)}건 / 결측 {missing}건")

# --- 엑셀로 깔끔하게 정리 ---
wb = Workbook()
ws = wb.active
ws.title = "상품 데이터"
headers = ["상품명", "가격", "평점(5점)", "재고 상태"]
ws.append(headers)

head_fill = PatternFill("solid", fgColor="2F5496")
head_font = Font(color="FFFFFF", bold=True, size=11)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for c in ws[1]:
    c.fill = head_fill
    c.font = head_font
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

for row in rows:
    ws.append(row)

for r_i, row in enumerate(ws.iter_rows(min_row=2), start=2):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(vertical="center")
    if r_i % 2 == 0:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="F2F5FB")

widths = [52, 12, 12, 16]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + i)].width = w
ws.freeze_panes = "A2"

out = "/private/tmp/claude-501/-Users-seojeong-won-GEMMA-4/265aacd5-b7ed-4f2a-9d79-2deb3e317e1f/scratchpad/portfolio_result.xlsx"
wb.save(out)
print(f"엑셀 저장: {out}")
